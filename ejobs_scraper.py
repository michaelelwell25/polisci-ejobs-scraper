"""
eJobs Scraper & Job Market Tracker
Parses APSA eJobs PDFs and MPSA eJobs CSVs, filters/scores listings,
and outputs a combined tracking spreadsheet.

Usage:
    python ejobs_scraper.py --apsa path/to/apsa.pdf --mpsa path/to/mpsa.csv -o output.xlsx
    python ejobs_scraper.py --apsa path/to/apsa.pdf -o output.xlsx
    python ejobs_scraper.py --mpsa path/to/mpsa.csv -o output.xlsx
    python ejobs_scraper.py --config my_profile.yaml --apsa apsa.pdf -o output.xlsx
"""

import argparse
import csv
import re
import sys
from dataclasses import dataclass, field, asdict
from pathlib import Path

try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Profile & Filter Config ─────────────────────────────────────────────────
# Defaults — overridden by profile_config.yaml if present

RESEARCH_KEYWORDS = [
    "race", "racial", "ethnicity", "ethnic", "identity", "identities",
    "lgbtq", "lgbt", "queer", "sexual orientation", "gender identity",
    "linked fate", "group consciousness", "solidarity",
    "political behavior", "political psychology", "public opinion",
    "voting behavior", "political attitudes", "political participation",
    "inequality", "marginalized", "minority politics", "minority",
    "intersectionality", "discrimination", "prejudice",
    "representation", "descriptive representation",
    "survey", "quantitative", "experimental", "causal inference",
]

SUBFIELD_MATCH = [
    "american politics", "american government", "methodology",
    "open field", "open", "political behavior", "political psychology",
    "public opinion",
]

SUBFIELD_EXCLUDE = []

CC_PATTERNS = [
    r"community college", r"community & technical",
    r"\bcc\b", r"two-year",
]

RANK_TT_PRIORITY = [
    "tt: assistant", "tt: open rank", "tt: assistant/associate",
    "assistant professor", "open rank",
]
RANK_VAP = [
    "visiting", "vap", "ntt: visiting",
]
RANK_EXCLUDE = [
    "dean", "director", "executive", "chair", "endowed",
    "postdoc", "post-doc", "research associate",
    "lecturer", "instructor", "ntt: lecturer", "ntt: teaching",
    "professor of practice",
]

ALLOWED_COUNTRIES = ["united states", "usa", "us", "canada", "canadian"]
EXCLUDE_CC = True
ALLOW_INTERNATIONAL = False


def load_config(config_path: str | None):
    """Load profile_config.yaml and override globals."""
    global RESEARCH_KEYWORDS, SUBFIELD_MATCH, SUBFIELD_EXCLUDE, RANK_TT_PRIORITY
    global RANK_VAP, RANK_EXCLUDE, ALLOWED_COUNTRIES, EXCLUDE_CC, ALLOW_INTERNATIONAL

    if config_path is None:
        # Try default location next to script
        default = Path(__file__).parent / "profile_config.yaml"
        if default.exists():
            config_path = str(default)
        else:
            return

    if not HAS_YAML:
        print("WARNING: PyYAML not installed, using default config. pip install pyyaml")
        return

    with open(config_path) as f:
        cfg = yaml.safe_load(f)

    if not cfg:
        return

    if "research_keywords" in cfg:
        RESEARCH_KEYWORDS = [kw.lower() for kw in cfg["research_keywords"]]
    if "subfield_match" in cfg:
        SUBFIELD_MATCH = [s.lower() for s in cfg["subfield_match"]]
    if "subfield_exclude" in cfg:
        SUBFIELD_EXCLUDE = [s.lower() for s in cfg["subfield_exclude"]]
    if "rank_tiers" in cfg:
        rt = cfg["rank_tiers"]
        if "priority" in rt:
            RANK_TT_PRIORITY = [r.lower() for r in rt["priority"]]
        if "lower_priority" in rt:
            RANK_VAP = [r.lower() for r in rt["lower_priority"]]
        if "exclude" in rt:
            RANK_EXCLUDE = [r.lower() for r in rt["exclude"]]
    if "locations" in cfg:
        loc = cfg["locations"]
        ALLOW_INTERNATIONAL = loc.get("international", False)
        EXCLUDE_CC = loc.get("exclude_community_colleges", True)
        if "allowed_countries" in loc:
            ALLOWED_COUNTRIES = ["united states", "usa", "us"] + \
                [c.lower() for c in loc["allowed_countries"]]

    print(f"Loaded config: {config_path}")

# US states + DC + territories for detecting US positions
US_STATES = {
    "alabama","alaska","arizona","arkansas","california","colorado",
    "connecticut","delaware","florida","georgia","hawaii","idaho",
    "illinois","indiana","iowa","kansas","kentucky","louisiana","maine",
    "maryland","massachusetts","michigan","minnesota","mississippi",
    "missouri","montana","nebraska","nevada","new hampshire","new jersey",
    "new mexico","new york","north carolina","north dakota","ohio",
    "oklahoma","oregon","pennsylvania","rhode island","south carolina",
    "south dakota","tennessee","texas","utah","vermont","virginia",
    "washington","west virginia","wisconsin","wyoming",
    "district of columbia",
}

CANADIAN_PROVINCES = {
    "ontario","quebec","british columbia","alberta","manitoba",
    "saskatchewan","nova scotia","new brunswick",
    "prince edward island","newfoundland","yukon",
    "northwest territories","nunavut",
}


@dataclass
class Job:
    source: str = ""
    institution: str = ""
    position: str = ""
    subfield: str = ""
    specialization: str = ""
    rank_raw: str = ""
    region: str = ""
    state: str = ""
    deadline: str = ""
    date_posted: str = ""
    salary: str = ""
    url: str = ""
    ejobs_id: str = ""
    start_date: str = ""
    tier: str = ""           # Priority, Lower Priority, Excluded
    match_score: int = 0     # keyword match count
    match_keywords: str = ""
    notes: str = ""
    full_text: str = ""


# ── Parsing: MPSA CSV ────────────────────────────────────────────────────────

def parse_mpsa_csv(path: str) -> list[Job]:
    jobs = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            j = Job(
                source="MPSA",
                institution=row.get("Institution", "").strip(),
                position=row.get("Position", "").strip(),
                subfield=row.get("Subfield", "").strip(),
                specialization=row.get("Specialization", "").strip(),
                rank_raw=row.get("Position Rank", "").strip(),
                region=row.get("Region", "").strip(),
                state=row.get("State", "").strip(),
                deadline=row.get("Application Deadline", "").strip(),
                date_posted=row.get("Date Added", "").strip(),
                full_text=str(row),
            )
            jobs.append(j)
    return jobs


# ── Parsing: APSA PDF ────────────────────────────────────────────────────────

SECTION_HEADERS = [
    "ADMINISTRATION",
    "AMERICAN GOVERNMENT AND POLITICS",
    "COMPARATIVE POLITICS",
    "INTERNATIONAL RELATIONS",
    "METHODOLOGY",
    "NON-ACADEMIC",
    "OPEN",
    "OTHER",
    "POLITICAL THEORY",
    "PUBLIC ADMINISTRATION",
    "PUBLIC LAW",
    "PUBLIC POLICY",
]

def _extract_field(text: str, label: str) -> str:
    pattern = rf'{label}\s*:\s*(.+?)(?:\n|$)'
    m = re.search(pattern, text, re.IGNORECASE)
    return m.group(1).strip() if m else ""

def _extract_urls(text: str) -> list[str]:
    return re.findall(r'https?://[^\s,;)]+', text)

def parse_apsa_pdf(path: str) -> list[Job]:
    if pdfplumber is None:
        print("ERROR: pdfplumber required for PDF parsing. pip install pdfplumber")
        sys.exit(1)

    pdf = pdfplumber.open(path)
    full_text = ""
    for page in pdf.pages:
        t = page.extract_text()
        if t:
            full_text += t + "\n"

    # Remove page headers
    full_text = re.sub(r'Political Science Jobs\s*(for\s*)?[A-Z][a-z]+ \d{4}\n?', '', full_text)
    full_text = re.sub(r'Political Science Jobs\s*[A-Z][a-z]+ \d{4}\n?', '', full_text)

    # Split into section blocks by section headers
    section_pattern = r'\n(' + '|'.join(re.escape(h) for h in SECTION_HEADERS) + r')\s*\n'
    parts = re.split(section_pattern, full_text)

    current_section = "Unknown"
    section_texts = []
    i = 0
    while i < len(parts):
        chunk = parts[i].strip()
        if chunk.upper() in [h.upper() for h in SECTION_HEADERS]:
            current_section = chunk.title()
            i += 1
            if i < len(parts):
                section_texts.append((current_section, parts[i]))
                i += 1
        else:
            i += 1

    jobs = []
    for section, text in section_texts:
        # Split listings by eJobs ID markers — each listing ends with eJobsID:NNNNN
        listings = re.split(r'(eJobsID:\s*\d+)', text)

        buf = ""
        for chunk in listings:
            buf += chunk
            if re.match(r'eJobsID:\s*\d+', chunk.strip()):
                j = _parse_single_apsa_listing(buf, section)
                if j:
                    jobs.append(j)
                buf = ""
        # Leftover without eJobsID — try to parse anyway
        if buf.strip() and len(buf.strip()) > 100:
            j = _parse_single_apsa_listing(buf, section)
            if j:
                jobs.append(j)

    return jobs


def _parse_single_apsa_listing(text: str, section: str) -> Job | None:
    text = text.strip()
    if len(text) < 50:
        return None

    # Extract structured fields
    ejobs_id = ""
    m = re.search(r'eJobsID:\s*(\d+)', text)
    if m:
        ejobs_id = m.group(1)

    rank = _extract_field(text, "Rank")
    subfields = _extract_field(text, r"Subfield\(s\)")
    specializations = _extract_field(text, "Specializations?")
    start_date = _extract_field(text, "StartDate")
    deadline = _extract_field(text, "ApplicationDeadline")
    date_posted = _extract_field(text, "DatePosted")
    salary = _extract_field(text, "Salary")

    # Try to extract institution — usually near the top
    # Look for university/college name patterns in first few lines
    lines = text.split("\n")
    institution = ""
    position_title = ""

    # Rank line often contains the position title
    if rank:
        position_title = rank

    # Institution: look for "University", "College", "Institute" in early lines
    inst_keywords = ["university", "college", "institute", "polytechnic", "academy"]
    for line in lines[:10]:
        line_clean = line.strip()
        # De-run PDF text: insert spaces before caps
        line_spaced = re.sub(r'([a-z])([A-Z])', r'\1 \2', line_clean)
        line_spaced = re.sub(r'(\w)(University|College|Institute)', r'\1 \2', line_spaced)
        lc = line_spaced.lower()
        if not any(kw in lc for kw in inst_keywords):
            continue
        if len(line_spaced) > 70:
            continue
        # Skip lines that are clearly descriptions, not names
        if any(skip in lc for skip in [
            "please", "must", "should", "will", "the department",
            "the school", "the successful", "for ", "across", "but ",
            "including", "review of", "this position", "invites",
            "seeks", "applications", "salary", "engage", "internship"]):
            continue
        # Clean: remove "Start Date:..." or "Open Date:..." prefixes
        line_spaced = re.sub(r'^(Start Date|Open Date|Salary\s*Range\s*or\s*Pay\s*Grade)[:\s].{5,30}\s+', '', line_spaced)
        # Remove trailing junk after institution name
        line_spaced = re.sub(r'\s+(journals|Salary|acrossall|convening).*$', '', line_spaced, flags=re.IGNORECASE)
        if len(line_spaced.strip()) < 5:
            continue
        institution = line_spaced.strip()
        break

    urls = _extract_urls(text)
    url = urls[0] if urls else ""

    # Clean up run-together fields from PDF extraction
    deadline = _clean_pdf_field(deadline)
    date_posted = _clean_pdf_field(date_posted)
    start_date = _clean_pdf_field(start_date)
    salary = _clean_pdf_field(salary)

    return Job(
        source="APSA",
        institution=institution,
        position=position_title,
        subfield=section + (f" / {_clean_pdf_field(subfields)}" if subfields else ""),
        specialization=specializations,
        rank_raw=rank,
        deadline=deadline,
        date_posted=date_posted,
        salary=salary,
        url=url,
        ejobs_id=ejobs_id,
        start_date=start_date,
        full_text=text[:2000],
    )


def _clean_pdf_field(val: str) -> str:
    if not val:
        return val
    # Insert spaces before capitals in camelCase runons
    val = re.sub(r'([a-z])([A-Z])', r'\1 \2', val)
    # Insert space between digit and letter: "26March2026" -> "26 March 2026"
    val = re.sub(r'(\d)([A-Za-z])', r'\1 \2', val)
    val = re.sub(r'([A-Za-z])(\d)', r'\1 \2', val)
    # Common normalizations
    val = val.replace("Open until Filled", "Open until filled")
    val = val.replace("Openuntil Filled", "Open until filled")
    # Truncate at first sentence-like junk
    m = re.match(r'^(.{3,60}?)(?:\.\s*[A-Z]|The[su ]|Please|Review|Application)', val)
    if m:
        val = m.group(1).strip().rstrip(".")
    # For dates, keep only the date-like portion
    dm = re.match(r'^(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})', val)
    if dm:
        val = dm.group(1)
    return val.strip()


# ── Filtering & Scoring ─────────────────────────────────────────────────────

def _is_cc(job: Job) -> bool:
    if not EXCLUDE_CC:
        return False
    blob = f"{job.institution} {job.position} {job.full_text}".lower()
    return any(re.search(p, blob) for p in CC_PATTERNS)

def _is_allowed_location(job: Job) -> bool:
    blob = f"{job.institution} {job.state} {job.region} {job.full_text}".lower()
    # If we have a state field and it matches US states, it's fine
    if job.state.lower().strip() in US_STATES:
        return True
    # Check for Canadian provinces
    if job.state.lower().strip() in CANADIAN_PROVINCES:
        return True
    # Check blob for Canada references
    if any(p in blob for p in ["canada", "canadian", "ontario", "quebec",
                                "british columbia", "alberta", "toronto",
                                "montreal", "vancouver", "ottawa"]):
        return True
    # MPSA has region field; if it's a known US region, fine
    if job.region.lower() in ["south", "northeast", "midwest", "west"]:
        return True
    # For APSA, check for non-US indicators
    non_us = ["united kingdom", "uk ", "oxford", "cambridge", "london",
              "australia", "germany", "france", "switzerland", "netherlands",
              "singapore", "hong kong", "japan", "korea", "china", "taiwan",
              "israel", "uae", "qatar", "saudi", "europe", "european",
              "afrika", "africa", "india", "brazil", "mexico", "chile",
              "new zealand", "ireland", "scotland", "wales", "denmark",
              "sweden", "norway", "finland", "italy", "spain", "portugal",
              "austria", "belgium", "czech", "poland", "hungary", "turkey",
              "gbp", "eur ", "£"]
    # Check institution name specifically for non-US indicators
    # Also check raw full_text first ~200 chars for institution context
    inst_lower = re.sub(r'([a-z])([A-Z])', r'\1 \2', job.institution).lower()
    text_head = re.sub(r'([a-z])([A-Z])', r'\1 \2', job.full_text[:300]).lower()
    if any(kw in inst_lower for kw in non_us):
        return False
    if any(kw in text_head for kw in non_us):
        return False
    # Also check full text but only flag if institution itself seems foreign
    if not job.state and not job.region:
        # No state/region info — check if full text has strong non-US signals
        if any(kw in blob for kw in non_us):
            # If no US/Canada signal at all, exclude
            us_signals = list(US_STATES) + list(CANADIAN_PROVINCES) + [
                "united states", "usa", "u.s.", "america"]
            if not any(s in blob for s in us_signals):
                return False
    return True

def _classify_rank(job: Job) -> str:
    blob = f"{job.rank_raw} {job.position}".lower()
    # Check exclusions first
    for pat in RANK_EXCLUDE:
        if pat in blob:
            # But don't exclude if it also says "assistant professor"
            if not any(tt in blob for tt in ["assistant professor", "open rank", "tt:"]):
                return "Excluded"
    # TT priority
    for pat in RANK_TT_PRIORITY:
        if pat in blob:
            return "Priority"
    # Check for tenure-track signals
    if "tenure" in blob and ("track" in blob or "stream" in blob):
        return "Priority"
    # VAP
    for pat in RANK_VAP:
        if pat in blob:
            return "Lower Priority"
    return "Excluded"

def _score_keywords(job: Job) -> tuple[int, list[str]]:
    blob = f"{job.position} {job.subfield} {job.specialization} {job.full_text}".lower()
    matched = []
    for kw in RESEARCH_KEYWORDS:
        if kw in blob and kw not in matched:
            matched.append(kw)
    return len(matched), matched

def _matches_subfield(job: Job) -> bool:
    blob = f"{job.subfield} {job.specialization}".lower()
    if SUBFIELD_EXCLUDE and any(sf in blob for sf in SUBFIELD_EXCLUDE):
        return False
    return any(sf in blob for sf in SUBFIELD_MATCH)

def filter_and_score(jobs: list[Job]) -> list[Job]:
    kept = []
    for j in jobs:
        if _is_cc(j):
            continue
        if not _is_allowed_location(j):
            continue
        tier = _classify_rank(j)
        if tier == "Excluded":
            continue
        score, kws = _score_keywords(j)
        # Filter by subfield — but let keyword-matched listings through
        if not _matches_subfield(j) and score == 0:
            continue
        j.tier = tier
        j.match_score = score
        j.match_keywords = ", ".join(kws)
        kept.append(j)

    # Sort: Priority first, then by match score descending
    tier_order = {"Priority": 0, "Lower Priority": 1}
    kept.sort(key=lambda j: (tier_order.get(j.tier, 2), -j.match_score))
    return kept


# ── Output ───────────────────────────────────────────────────────────────────

OUTPUT_COLS = [
    ("Source", "source"),
    ("Tier", "tier"),
    ("Match Score", "match_score"),
    ("Institution", "institution"),
    ("Position", "position"),
    ("Subfield", "subfield"),
    ("Specialization", "specialization"),
    ("Rank", "rank_raw"),
    ("Region", "region"),
    ("State", "state"),
    ("Deadline", "deadline"),
    ("Start Date", "start_date"),
    ("Date Posted", "date_posted"),
    ("Salary", "salary"),
    ("URL", "url"),
    ("eJobs ID", "ejobs_id"),
    ("Matched Keywords", "match_keywords"),
    ("Applied?", "notes"),  # blank for user to fill in
    ("Status", "notes"),    # blank for user to fill in
]

def write_xlsx(jobs: list[Job], path: str):
    if not HAS_OPENPYXL:
        print("openpyxl not installed, falling back to CSV")
        write_csv(jobs, path.replace(".xlsx", ".csv"))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    priority_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    vap_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write headers
    headers = [c[0] for c in OUTPUT_COLS]
    # Replace duplicate "notes" columns with user-facing blanks
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Write data
    for row_idx, job in enumerate(jobs, 2):
        for col_idx, (_, attr) in enumerate(OUTPUT_COLS, 1):
            # "Applied?" and "Status" columns start blank
            if OUTPUT_COLS[col_idx-1][0] in ("Applied?", "Status"):
                val = ""
            else:
                val = getattr(job, attr, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Row coloring by tier
        fill = priority_fill if job.tier == "Priority" else vap_fill
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Column widths
    col_widths = {
        "Source": 8, "Tier": 14, "Match Score": 10, "Institution": 30,
        "Position": 35, "Subfield": 22, "Specialization": 30, "Rank": 25,
        "Region": 10, "State": 14, "Deadline": 14, "Start Date": 12,
        "Date Posted": 12, "Salary": 16, "URL": 35, "eJobs ID": 10,
        "Matched Keywords": 30, "Applied?": 10, "Status": 12,
    }
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 15)

    # Freeze top row + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    print(f"Wrote {len(jobs)} jobs to {path}")

def write_csv(jobs: list[Job], path: str):
    headers = [c[0] for c in OUTPUT_COLS]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for job in jobs:
            row = []
            for _, attr in OUTPUT_COLS:
                if OUTPUT_COLS[len(row)][0] in ("Applied?", "Status"):
                    row.append("")
                else:
                    row.append(getattr(job, attr, ""))
            w.writerow(row)
    print(f"Wrote {len(jobs)} jobs to {path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="eJobs Scraper & Job Market Tracker")
    parser.add_argument("--apsa", help="Path to APSA eJobs PDF")
    parser.add_argument("--mpsa", help="Path to MPSA eJobs CSV")
    parser.add_argument("--config", help="Path to profile_config.yaml (default: auto-detect)")
    parser.add_argument("-o", "--output", default="job_tracker.xlsx",
                        help="Output file path (xlsx or csv)")
    args = parser.parse_args()

    load_config(args.config)

    if not args.apsa and not args.mpsa:
        print("Provide at least one of --apsa or --mpsa")
        sys.exit(1)

    all_jobs = []

    if args.apsa:
        print(f"Parsing APSA PDF: {args.apsa}")
        apsa_jobs = parse_apsa_pdf(args.apsa)
        print(f"  Found {len(apsa_jobs)} raw listings")
        all_jobs.extend(apsa_jobs)

    if args.mpsa:
        print(f"Parsing MPSA CSV: {args.mpsa}")
        mpsa_jobs = parse_mpsa_csv(args.mpsa)
        print(f"  Found {len(mpsa_jobs)} raw listings")
        all_jobs.extend(mpsa_jobs)

    # Deduplicate by institution + position (fuzzy)
    seen = set()
    deduped = []
    for j in all_jobs:
        key = (j.institution.lower().strip()[:40], j.position.lower().strip()[:40])
        if key not in seen:
            seen.add(key)
            deduped.append(j)
        else:
            # Keep the one with more info
            pass
    all_jobs = deduped

    print(f"After dedup: {len(all_jobs)} listings")
    filtered = filter_and_score(all_jobs)
    print(f"After filtering: {len(filtered)} listings")
    print(f"  Priority (TT): {sum(1 for j in filtered if j.tier == 'Priority')}")
    print(f"  Lower Priority (VAP): {sum(1 for j in filtered if j.tier == 'Lower Priority')}")

    if args.output.endswith(".csv"):
        write_csv(filtered, args.output)
    else:
        write_xlsx(filtered, args.output)


if __name__ == "__main__":
    main()
